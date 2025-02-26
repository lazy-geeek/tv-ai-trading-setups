import base64
import os
import json
import time
from datetime import datetime
from glob import glob

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openai import OpenAI
from decouple import config
from tqdm import tqdm
from playwright.sync_api import sync_playwright

from llm_prompts import (
    SUMMARY_SYSTEM_PROMPT,
    TRADING_SYSTEM_PROMPT,
    TRADING_USER_PROMPT,
)
from helper_func import (
    print_status,
    save_trading_setup_to_file,
    get_symbol_directory,
    get_download_directory,
    get_trading_setups_directory,
    clear_download_directory,
)

# Configuration
endpoint_url = config("ENDPOINT_URL")
website_url = config("WEBSITE_URL")
tf_reload_timeout = int(config("TF_RELOAD_TIMEOUT"))
chart_reload_timeout = int(config("CHART_RELOAD_TIMEOUT"))
timeframes = json.loads(config("TIMEFRAMES"))
symbols = json.loads(config("SYMBOLS"))

openai_model = config("OPENAI_MODEL")
openai_api_key = config("OPENAI_API_KEY")
openai_base_url = "https://api.openai.com/v1/"

gemini_api_key = config("GEMINI_API_KEY")
gemini_model = config("GEMINI_MODEL")
gemini_base_url = "https://generativelanguage.googleapis.com/v1beta/openai/"

openrouter_api_key = config("OPENROUTER_API_KEY")
openrouter_base_url = "https://openrouter.ai/api/v1"

anthropic_model = config("ANTHROPIC_MODEL")
summary_model = config("SUMMARY_MODEL")


# --- Screenshot Functionality ---
def take_screenshots_for_symbol(symbol):
    print_status(f"Taking Tradingview screenshots for {symbol}...")

    with sync_playwright() as p:
        # Connect to the existing browser instance
        browser = p.chromium.connect_over_cdp(endpoint_url)

        # Get the default browser context (this will use the existing user profile)
        context = browser.contexts[0]

        # Create a new page in the existing context
        page = context.new_page()
        page.set_default_timeout(10000)

        # Navigate to a website (it should already be logged in)
        page.goto(website_url)

        download_directory = get_symbol_directory(symbol)

        page.keyboard.type(symbol)
        page.wait_for_timeout(tf_reload_timeout)
        page.keyboard.press("Enter")
        page.wait_for_timeout(tf_reload_timeout)

        # Loop through timeframes for the current symbol with a progress bar
        for timeframe in tqdm(timeframes, desc=f"Timeframes for {symbol}", leave=False):
            # Change timeframe by typing
            page.keyboard.type(timeframe)
            page.wait_for_timeout(tf_reload_timeout)
            page.keyboard.press("Enter")
            page.wait_for_timeout(chart_reload_timeout)

            # Download screenshot
            with page.expect_download() as download_info:
                page.keyboard.press("Control+Alt+S")

            download = download_info.value

            # Wait for the download process to complete and save the downloaded file in specified path
            download.save_as(
                os.path.join(download_directory, download.suggested_filename)
            )

        # Close page
        page.close()

    print_status(f"Screenshots completed for {symbol}!")


# --- Trading Setup Generation Functionality ---
def generate_setups_for_symbol(symbol):
    print_status(f"Generating trading setups for {symbol}...")

    symbol_dir = get_symbol_directory(symbol)

    # Get list of all files in the download directory
    screenshot_files = glob(os.path.join(symbol_dir, "*.png"))
    if not screenshot_files:
        print_status(
            f"Warning: No screenshots found in download directory: {symbol_dir}"
        )
        return

    with tqdm(total=3, desc=f"Generating setups {symbol}", leave=False) as pbar:
        chatgpt_setup = get_chatgpt_trading_setup(screenshot_files)
        save_trading_setup_to_file(symbol, chatgpt_setup, openai_model)
        pbar.update(1)

        gemini_setup = get_gemini_trading_setup(screenshot_files)
        save_trading_setup_to_file(symbol, gemini_setup, gemini_model)
        pbar.update(1)

        anthropic_setup = get_openai_trading_setup(
            openrouter_api_key,
            openrouter_base_url,
            anthropic_model,
            screenshot_files,
        )
        save_trading_setup_to_file(symbol, anthropic_setup, anthropic_model)
        pbar.update(1)


def openai_message_content(screenshot_files):
    system_role = {"role": "system", "content": TRADING_SYSTEM_PROMPT}
    user_role = {
        "role": "user",
        "content": [{"type": "text", "text": TRADING_USER_PROMPT}],
    }

    # Add images to the user message content
    images_processed = 0
    for screenshot in screenshot_files:
        try:
            with open(screenshot, "rb") as file:
                base64_image = base64.b64encode(file.read()).decode("utf-8")
                user_role["content"].append(
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{base64_image}"},
                    }
                )
                images_processed += 1
        except Exception as e:
            print_status(f"Error processing image {screenshot}: {e}")
            continue

    if images_processed == 0:
        print_status("No valid images were processed. Cannot proceed with analysis.")
        return None

    messages = [system_role, user_role]

    return messages


def get_openai_trading_setup(openai_api_key, openai_base_url, model, screenshot_files):
    try:
        client = OpenAI(api_key=openai_api_key, base_url=openai_base_url)
        messages = openai_message_content(screenshot_files)

        response = client.chat.completions.create(model=model, messages=messages)

        return response.choices[0].message.content

    except Exception as e:
        print_status(f"Error: {e}")
        return None


def get_chatgpt_trading_setup(screenshot_files):
    return get_openai_trading_setup(
        openai_api_key, openai_base_url, openai_model, screenshot_files
    )


def get_gemini_trading_setup(screenshot_files):
    return get_openai_trading_setup(
        gemini_api_key, gemini_base_url, gemini_model, screenshot_files
    )


# --- Summary Generation Functionality ---
def summarize_setups_for_symbol(symbol):
    print_status(f"Summarizing trading setups for {symbol}...")

    client = OpenAI(api_key=openrouter_api_key, base_url=openrouter_base_url)

    directory = get_trading_setups_directory(symbol)
    files = [f for f in os.listdir(directory) if f.lower().endswith(".txt")]
    summaries = []

    for filename in tqdm(files, desc=f"Processing files for {symbol}", leave=False):
        file_path = os.path.join(directory, filename)
        with open(file_path, "r", encoding="utf-8") as f:
            text = f.read()

        client = OpenAI(api_key=openrouter_api_key, base_url=openrouter_base_url)

        response = client.chat.completions.create(
            model=summary_model,
            messages=[
                {"role": "system", "content": SUMMARY_SYSTEM_PROMPT},
                {"role": "user", "content": text},
            ],
        )

        if isinstance(response, str):
            print(f"OpenRouter API Error: {response}")
            summary_text = ""
        else:
            summary_text = response.choices[0].message.content

        # Remove code block delimiters
        summary_text = summary_text.replace("```json", "").replace("```", "")
        try:
            summary = json.loads(summary_text)
        except json.JSONDecodeError as e:
            print(f"JSONDecodeError: {e}")  # Add logging
            summary = {
                "direction": None,
                "entry": None,
                "stop_loss": None,
                "take_profit": None,
                "rrr": None,
                "stop_loss_pips": None,
                "take_profit_pips": None,
            }
        summary["filename"] = os.path.splitext(filename)[0]
        summaries.append(summary)
    return summaries


# --- Excel Export Functionality ---
def save_summaries_to_excel_for_symbol(symbol, summaries):
    print_status(f"Saving summaries to Excel for {symbol}...")

    workbook = Workbook()
    # Remove the default sheet created by Workbook
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheet = workbook.create_sheet(title=symbol)
    sheet.append(
        [
            "Filename",
            "Direction",
            "Entry",
            "Stop Loss",
            "Take Profit",
            "RRR",
            "Stop Loss Pips",
            "Take Profit Pips",
        ]
    )
    for summary in summaries:
        sheet.append(
            [
                summary.get("filename"),
                summary.get("direction"),
                summary.get("entry"),
                summary.get("stop_loss"),
                summary.get("take_profit"),
                summary.get("rrr"),
                summary.get("stop_loss_pips"),
                summary.get("take_profit_pips"),
            ]
        )

    # Apply formatting to the sheet
    # Bold header row cells
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    # Bold first column cells
    for row in sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)
    # Auto-size columns
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width

    download_directory = get_download_directory()
    timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M")
    filename = f"{symbol}_trading_summaries-{timestamp}.xlsx"
    save_path = os.path.join(download_directory, filename)
    if os.path.exists(save_path):
        os.remove(save_path)
    workbook.save(save_path)
    print_status(f"Summaries written to {save_path}")


# --- Main Function ---
def main():
    print_status("Starting main process...")

    # Delete all files from download directory
    clear_download_directory()

    for symbol in tqdm(symbols, desc="Processing Symbols"):
        print_status(f"Starting process for symbol: {symbol}")

        take_screenshots_for_symbol(symbol)
        generate_setups_for_symbol(symbol)
        summaries = summarize_setups_for_symbol(symbol)
        save_summaries_to_excel_for_symbol(symbol, summaries)

        print_status(f"Completed process for symbol: {symbol}")

    print_status("Main process completed!")


if __name__ == "__main__":
    main()
