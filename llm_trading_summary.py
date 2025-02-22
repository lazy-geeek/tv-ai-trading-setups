import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openai import OpenAI
from decouple import config
import time
from datetime import datetime

from helper_func import (
    get_download_directory,
    get_trading_setups_directory,
    print_status,
)
from llm_prompts import SUMMARY_SYSTEM_PROMPT
from tqdm import tqdm

symbols = json.loads(config("SYMBOLS"))
download_directory = get_download_directory()

model = config("SUMMARY_MODEL")
api_key = config("OPENROUTER_API_KEY")
base_url = "https://openrouter.ai/api/v1"


def summarize_trading_setups():

    print_status("Summarizing trading setups to Excel...")

    workbook = Workbook()
    # Remove the default sheet created by Workbook
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    client = OpenAI(api_key=api_key, base_url=base_url)

    for symbol in tqdm(symbols, desc="Processing symbols"):
        directory = get_trading_setups_directory(symbol)
        files = [f for f in os.listdir(directory) if f.lower().endswith(".txt")]
        directions = set()
        summaries = []
        for filename in tqdm(files, desc=f"Processing files for {symbol}", leave=False):
            file_path = os.path.join(directory, filename)
            with open(file_path, "r", encoding="utf-8") as f:
                text = f.read()
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": SUMMARY_SYSTEM_PROMPT},
                    {"role": "user", "content": text},
                ],
            )
            if isinstance(response, str):
                print(f"OpenRouter API Error: {response}")
                summary_text = ""
            else:
                summary_text = response.choices[0].message.content
            # Remove code block delimiters
            summary_text = summary_text.replace("```json", "").replace("```", "")
            try:
                summary = json.loads(summary_text)
            except json.JSONDecodeError as e:
                print(f"JSONDecodeError: {e}")  # Add logging
                summary = {
                    "direction": None,
                    "entry": None,
                    "stop_loss": None,
                    "take_profit": None,
                    "rrr": None,
                    "stop_loss_pips": None,
                    "take_profit_pips": None,
                }
            summary["filename"] = os.path.splitext(filename)[0]
            directions.add(summary.get("direction"))
            summaries.append(summary)

        if len(directions) == 1 and "None" not in directions:
            sheet = workbook.create_sheet(title=symbol)
            sheet.append(
                [
                    "Filename",
                    "Direction",
                    "Entry",
                    "Stop Loss",
                    "Take Profit",
                    "RRR",
                    "Stop Loss Pips",
                    "Take Profit Pips",
                ]
            )
            for summary in summaries:
                sheet.append(
                    [
                        summary.get("filename"),
                        summary.get("direction"),
                        summary.get("entry"),
                        summary.get("stop_loss"),
                        summary.get("take_profit"),
                        summary.get("rrr"),
                        summary.get("stop_loss_pips"),
                        summary.get("take_profit_pips"),
                    ]
                )

    if len(workbook.worksheets) == 0:
        print_status("No valid trading summaries found. Excel file not generated.")
        return

    # Apply formatting to all sheets after processing
    for sheet in workbook.worksheets:
        # Bold header row cells
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        # Bold first column cells
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True)
        # Auto-size columns
        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max_length + 2
            sheet.column_dimensions[col_letter].width = adjusted_width
    timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M")
    filename = f"trading_summaries-{timestamp}.xlsx"
    save_path = os.path.join(download_directory, filename)
    if os.path.exists(save_path):
        os.remove(save_path)
    workbook.save(save_path)
    print_status(f"Summaries written to {save_path}")
